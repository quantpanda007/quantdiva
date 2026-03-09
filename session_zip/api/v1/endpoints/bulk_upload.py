"""
Bulk upload endpoint — accepts Excel file with multiple FX forward deals,
prices each row, returns results as downloadable Excel.
"""

from __future__ import annotations

import io
import math
import logging
import tempfile
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

router = APIRouter()
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column mapping — maps business team's Excel headers to internal field names
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    # Transaction ref
    "transaction ref no": "transaction_ref",
    "transaction ref": "transaction_ref",
    "trans ref": "transaction_ref",
    "ref no": "transaction_ref",
    "ref": "transaction_ref",
    # Client
    "client name": "client_name",
    "client": "client_name",
    # Counterparties
    "counterparty a": "cpty_a",
    "counterparty b": "cpty_b",
    "counterparty": "cpty_a",
    "counterp arty a": "cpty_a",
    "counterp arty b": "cpty_b",
    # Contract type
    "buy contract": "buy_contract",
    "sell contract": "sell_contract",
    # Dates
    "transaction date / trade date": "trade_date",
    "transaction date": "trade_date",
    "trade date": "trade_date",
    "effective date": "effective_date",
    "valuation / reporting date": "reporting_date",
    "valuation date": "reporting_date",
    "reporting date": "reporting_date",
    "maturity date": "maturity_date",
    "maturity da": "maturity_date",
    "maturity": "maturity_date",
    # Instrument
    "type of contract": "contract_type",
    "type": "contract_type",
    # Market data
    "spot": "spot",
    "strike rate": "strike",
    "strike": "strike",
    # Currency
    "currency pair": "ccy_pair",
    "currency pa": "ccy_pair",
    "ccy pair": "ccy_pair",
    # Notionals
    "notional currency 1": "notional_1",
    "notional ccy 1": "notional_1",
    "notional 1": "notional_1",
    "notional": "notional_1",
    "notional ccy 2": "notional_2",
    "notional currency 2": "notional_2",
    # Direction
    "buy /sell": "direction_1",
    "buy/sell": "direction_1",
    # Rates (optional columns)
    "domestic rate": "domestic_rate",
    "foreign rate": "foreign_rate",
}


def _normalize_header(h: str) -> str:
    """Normalize header for matching."""
    return h.strip().lower().replace("\n", " ").replace("  ", " ")


def _parse_date(val) -> Optional[date]:
    """Parse date from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(val) -> float:
    """Parse number, handling Indian comma format (10,00,000)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# Parse uploaded Excel
# ---------------------------------------------------------------------------

def parse_upload(file_bytes: bytes) -> Tuple[List[Dict], List[Dict]]:
    """
    Parse the uploaded Excel file.
    Returns (deals, errors) where each deal is a dict with mapped fields.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(_normalize_header(str(cell.value or "")))

    # Map columns
    col_map = {}
    direction_cols = []
    for idx, h in enumerate(headers):
        matched = COLUMN_ALIASES.get(h)
        if matched:
            if matched == "direction_1":
                direction_cols.append(idx)
            else:
                col_map[matched] = idx
        # Handle duplicate "Buy /Sell" columns — first is CCY1, second is CCY2
    if len(direction_cols) >= 1:
        col_map["direction_1"] = direction_cols[0]
    if len(direction_cols) >= 2:
        col_map["direction_2"] = direction_cols[1]

    deals = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue  # skip empty rows

        def get(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        # Parse fields
        deal = {
            "row": row_idx,
            "transaction_ref": str(get("transaction_ref") or f"ROW-{row_idx}"),
            "client_name": str(get("client_name") or ""),
            "cpty_a": str(get("cpty_a") or ""),
            "cpty_b": str(get("cpty_b") or ""),
            "buy_contract": str(get("buy_contract") or ""),
            "sell_contract": str(get("sell_contract") or ""),
            "trade_date": _parse_date(get("trade_date")),
            "effective_date": _parse_date(get("effective_date")),
            "reporting_date": _parse_date(get("reporting_date")),
            "contract_type": str(get("contract_type") or "Forward"),
            "spot": _parse_number(get("spot")),
            "strike": _parse_number(get("strike")),
            "ccy_pair": str(get("ccy_pair") or "USDINR").replace("/", ""),
            "notional_1": _parse_number(get("notional_1")),
            "direction_1": str(get("direction_1") or "Sell"),
            "notional_2": _parse_number(get("notional_2")),
            "maturity_date": _parse_date(get("maturity_date")),
            "domestic_rate": _parse_number(get("domestic_rate")) if get("domestic_rate") else None,
            "foreign_rate": _parse_number(get("foreign_rate")) if get("foreign_rate") else None,
        }

        # Validate required fields
        errs = []
        if not deal["strike"]:
            errs.append("Missing strike rate")
        if not deal["notional_1"]:
            errs.append("Missing notional")
        if not deal["maturity_date"]:
            errs.append("Missing or invalid maturity date")
        if not deal["spot"]:
            errs.append("Missing spot rate")

        if errs:
            errors.append({"row": row_idx, "ref": deal["transaction_ref"], "errors": "; ".join(errs)})
        else:
            deals.append(deal)

    return deals, errors


# ---------------------------------------------------------------------------
# Price deals using existing PricingService
# ---------------------------------------------------------------------------

def price_deals(
    deals: List[Dict],
    default_domestic_rate: float,
    default_foreign_rate: float,
) -> Tuple[List[Dict], List[Dict]]:
    """
    Price each deal using the existing FX Forward pricing.
    Returns (results, pricing_errors).
    """
    from services.pricers.pricing_service import PricingService
    from api.v1.helpers import build_instrument_from_request, build_market_env_from_request
    from api.v1.schemas import InstrumentRequest, MarketDataRequest, UnderlyingData

    ps = PricingService()
    results = []
    pricing_errors = []

    for deal in deals:
        try:
            r_d = deal["domestic_rate"] if deal["domestic_rate"] is not None else default_domestic_rate
            r_f = deal["foreign_rate"] if deal["foreign_rate"] is not None else default_foreign_rate

            pricing_date = deal["reporting_date"] or date.today()

            # Direction mapping
            direction = "sell" if deal["direction_1"].lower() in ("sell", "s") else "buy"

            # Build instrument request
            inst_req = InstrumentRequest(
                type="fx_forward",
                params={
                    "ccy_pair": deal["ccy_pair"],
                    "strike": deal["strike"],
                    "delivery_date": deal["maturity_date"].isoformat(),
                    "notional": deal["notional_1"],
                    "direction": direction,
                }
            )

            # Build market data
            md_req = MarketDataRequest(
                pricing_date=pricing_date.isoformat(),
                underlyings={deal["ccy_pair"]: UnderlyingData(spot=deal["spot"], vol=0.06)},
                rate_curve=[{"tenor": "1Y", "rate": r_d}],
                foreign_rate=r_f,
            )

            instrument = build_instrument_from_request(inst_req)
            underlying = getattr(instrument, "ccy_pair", "")
            market_env = build_market_env_from_request(md_req, underlying=underlying)

            result = ps.price(
                instrument, market_env,
                model_type="black_scholes",
                engine_type="analytic",
            )

            # Compute forward rate and DF for output
            from datetime import date as dt
            day_count_val = (deal["maturity_date"] - pricing_date).days / 365.0
            if day_count_val > 0:
                fwd_rate = deal["spot"] * math.exp((r_d - r_f) * day_count_val)
                disc_factor = math.exp(-r_d * day_count_val)
            else:
                fwd_rate = deal["spot"]
                disc_factor = 1.0

            # Long/short term classification
            months_to_maturity = (deal["maturity_date"] - pricing_date).days / 30.0
            long_term = result.npv if months_to_maturity > 12 else 0.0
            short_term = result.npv if months_to_maturity <= 12 else 0.0

            results.append({
                **deal,
                "npv": result.npv,
                "forward_rate": fwd_rate,
                "discount_factor": disc_factor,
                "long_term": long_term,
                "short_term": short_term,
                "domestic_rate_used": r_d,
                "foreign_rate_used": r_f,
                "pricing_date": pricing_date,
                "status": "OK",
                "elapsed_ms": result.diagnostics.get("elapsed_seconds", 0) * 1000 if result.diagnostics else 0,
            })

        except Exception as e:
            pricing_errors.append({
                "row": deal["row"],
                "ref": deal["transaction_ref"],
                "errors": str(e),
            })

    return results, pricing_errors


# ---------------------------------------------------------------------------
# Generate results Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1a2540")
HEADER_FONT = Font(bold=True, color="F59E0B", size=11)
OK_FONT = Font(color="10B981")
ERR_FONT = Font(color="EF4444")
NUM_FMT_MONEY = '#,##0'
NUM_FMT_RATE = '0.0000'
NUM_FMT_PCT = '0.00%'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="2E4A7A"),
)


def generate_results_excel(
    results: List[Dict],
    errors: List[Dict],
    upload_filename: str,
) -> bytes:
    """Generate formatted results Excel workbook."""
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 35

    summary_data = [
        ("Upload File", upload_filename),
        ("Total Deals Processed", len(results) + len(errors)),
        ("Priced Successfully", len(results)),
        ("Errors", len(errors)),
        ("", ""),
        ("Total NPV", sum(r["npv"] for r in results)),
        ("Total Long Term", sum(r["long_term"] for r in results)),
        ("Total Short Term", sum(r["short_term"] for r in results)),
        ("", ""),
        ("Pricing Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    if results:
        summary_data.insert(4, ("Pricing Date", str(results[0].get("pricing_date", ""))))

    for i, (label, val) in enumerate(summary_data, 1):
        ws_sum.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        cell = ws_sum.cell(row=i, column=2, value=val)
        if isinstance(val, float):
            cell.number_format = NUM_FMT_MONEY

    # --- Sheet 2: Deal Results ---
    ws_res = wb.create_sheet("Deal_Results")
    headers = [
        "Transaction Ref", "Client Name", "Cpty A", "Cpty B",
        "Ccy Pair", "Strike Rate", "Notional", "Direction",
        "Trade Date", "Effective Date", "Maturity Date", "Pricing Date",
        "Spot", "Domestic Rate", "Foreign Rate",
        "Forward Rate", "Discount Factor",
        "NPV", "Long Term", "Short Term", "Status",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws_res.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, r in enumerate(results, 2):
        data = [
            r["transaction_ref"],
            r["client_name"],
            r["cpty_a"],
            r["cpty_b"],
            r["ccy_pair"],
            r["strike"],
            r["notional_1"],
            r["direction_1"],
            r["trade_date"].isoformat() if r.get("trade_date") else "",
            r["effective_date"].isoformat() if r.get("effective_date") else "",
            r["maturity_date"].isoformat() if r.get("maturity_date") else "",
            str(r.get("pricing_date", "")),
            r["spot"],
            r["domestic_rate_used"],
            r["foreign_rate_used"],
            r["forward_rate"],
            r["discount_factor"],
            r["npv"],
            r["long_term"],
            r["short_term"],
            r["status"],
        ]
        for col, val in enumerate(data, 1):
            cell = ws_res.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            # Format numbers
            if col in (6, 13):  # strike, spot
                cell.number_format = NUM_FMT_RATE
            elif col in (7,):  # notional
                cell.number_format = NUM_FMT_MONEY
            elif col in (14, 15):  # rates
                cell.number_format = NUM_FMT_PCT
            elif col == 16:  # forward rate
                cell.number_format = NUM_FMT_RATE
            elif col == 17:  # discount factor
                cell.number_format = '0.000000'
            elif col in (18, 19, 20):  # NPV, long, short
                cell.number_format = NUM_FMT_MONEY
                cell.font = Font(color="10B981" if val >= 0 else "EF4444") if isinstance(val, (int, float)) else None
            elif col == 21:  # status
                cell.font = OK_FONT

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws_res.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(14, len(headers[col-1]) + 4)

    # --- Sheet 3: Errors (if any) ---
    if errors:
        ws_err = wb.create_sheet("Errors")
        err_headers = ["Row", "Transaction Ref", "Error"]
        for col, h in enumerate(err_headers, 1):
            cell = ws_err.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for i, e in enumerate(errors, 2):
            ws_err.cell(row=i, column=1, value=e["row"])
            ws_err.cell(row=i, column=2, value=e["ref"])
            ws_err.cell(row=i, column=3, value=e["errors"]).font = ERR_FONT
        ws_err.column_dimensions["A"].width = 8
        ws_err.column_dimensions["B"].width = 25
        ws_err.column_dimensions["C"].width = 60

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Template generator
# ---------------------------------------------------------------------------

def generate_template() -> bytes:
    """Generate a blank template Excel with correct headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FX_Forwards"

    headers = [
        "Transaction Ref no", "Client Name", "Counterparty A",
        "Buy Contract", "Counterparty B", "Sell Contract",
        "Transaction Date / Trade Date", "Effective Date",
        "Valuation / Reporting Date", "Type of Contract",
        "Spot", "Strike Rate", "Currency PaiR",
        "Notional Currency 1", "Buy /Sell",
        "Notional CCY 2", "Buy /Sell",
        "Maturity Date",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sample row
    sample = [
        "L01SFWD000130", "EXL India", "EXL India",
        "Buy", "BOA", "Sell",
        "07-11-2022", "07-11-2022",
        "31-03-2025", "Forward",
        85.47, 86.88, "USDINR",
        1000000, "Sell",
        8687750, "Buy",
        "29-07-2025",
    ]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(15, len(headers[col-1]) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/bulk-upload")
async def bulk_upload(
    file: UploadFile = File(...),
    domestic_rate: float = Query(0.065, description="Domestic rate (e.g. 0.065 for 6.5%)"),
    foreign_rate: float = Query(0.045, description="Foreign rate (e.g. 0.045 for 4.5%)"),
):
    """
    Upload Excel with FX forward deals, price all, return results Excel.

    Accepts the business team's standard format.
    Domestic/foreign rates are provided as query params (applied to all deals
    unless the Excel has rate columns per row).
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files accepted")

    try:
        contents = await file.read()
        deals, parse_errors = parse_upload(contents)

        if not deals and parse_errors:
            raise HTTPException(400, f"No valid deals found. Errors: {parse_errors}")

        results, pricing_errors = price_deals(deals, domestic_rate, foreign_rate)
        all_errors = parse_errors + pricing_errors

        result_bytes = generate_results_excel(results, all_errors, file.filename)

        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="Optima_BulkResults_{file.filename}"'
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Bulk upload failed")
        raise HTTPException(500, f"Bulk upload failed: {e}")


@router.get("/bulk-template")
async def download_template():
    """Download blank Excel template with correct column headers."""
    template_bytes = generate_template()
    return StreamingResponse(
        io.BytesIO(template_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Optima_BulkUpload_Template.xlsx"'
        },
    )
