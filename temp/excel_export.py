"""
Excel export for pricing results.

Generates professional Excel workbooks with:
  Sheet 1: Pricing Summary — inputs, market data, NPV, greeks
  Sheet 2: Greeks Detail — all sensitivities with bump details
  Sheet 3: MC Simulations (if Monte Carlo) — paths, payoffs, statistics
  Sheet 4: MC Random Numbers (if Monte Carlo) — raw random draws
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styles ───────────────────────────────────────────────────

NAVY = "1B2A4A"
ACCENT = "2E86AB"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(name="Arial", bold=True, color=NAVY, size=11)
VALUE_FONT = Font(name="Arial", size=11)
INPUT_FONT = Font(name="Arial", color="0000FF", size=11)  # Blue = inputs
NPV_FONT = Font(name="Arial", bold=True, color=ACCENT, size=14)
SECTION_FILL = PatternFill("solid", fgColor="EDF4F8")
SECTION_FONT = Font(name="Arial", bold=True, color=NAVY, size=12)
ALT_ROW_FILL = PatternFill("solid", fgColor="F5F5F5")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
NUMBER_FMT = '#,##0.0000'
DOLLAR_FMT = '$#,##0.0000'
PCT_FMT = '0.00%'
INT_FMT = '#,##0'


def _set_col_widths(ws, widths: Dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_label_value(ws, row, label, value, col_label=1, col_value=2,
                       val_font=None, val_fmt=None):
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = LABEL_FONT
    lc.alignment = Alignment(horizontal="right")

    vc = ws.cell(row=row, column=col_value, value=value)
    vc.font = val_font or VALUE_FONT
    if val_fmt:
        vc.number_format = val_fmt
    return row + 1


def _section_header(ws, row, text, col_span=4):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, col_span + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row + 1


# ── Main Export Function ─────────────────────────────────────

def generate_pricing_excel(
    instrument: Dict[str, Any],
    market_data: Dict[str, Any],
    model: str,
    engine: str,
    result: Dict[str, Any],
    greeks: Optional[Dict[str, Any]] = None,
    engine_params: Optional[Dict[str, Any]] = None,
    mc_data: Optional[Dict[str, Any]] = None,
) -> bytes:
    """Generate Excel workbook with full pricing details.

    Args:
        instrument: {"type": "...", "params": {...}}
        market_data: Market data dict
        model: Model name
        engine: Engine name
        result: Pricing result dict (npv, elapsed_ms, etc.)
        greeks: Greeks dict (delta, gamma, vega, etc.)
        engine_params: Engine-specific params (num_paths, etc.)
        mc_data: Monte Carlo simulation data (paths, payoffs, randoms)

    Returns:
        Excel file as bytes
    """
    wb = Workbook()

    # Sheet 1: Pricing Summary
    _build_summary_sheet(wb, instrument, market_data, model, engine,
                         result, greeks, engine_params)

    # Sheet 2: Greeks Detail
    if greeks:
        _build_greeks_sheet(wb, greeks, result)

    # Sheet 3 & 4: MC Simulation Data
    if mc_data:
        _build_mc_paths_sheet(wb, mc_data, instrument)
        if mc_data.get("random_numbers") is not None:
            _build_mc_randoms_sheet(wb, mc_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Sheet 1: Pricing Summary ────────────────────────────────

def _build_summary_sheet(wb, instrument, market_data, model, engine,
                         result, greeks, engine_params):
    ws = wb.active
    ws.title = "Pricing Summary"
    _set_col_widths(ws, {"A": 22, "B": 28, "C": 5, "D": 22, "E": 28})
    ws.sheet_properties.tabColor = NAVY

    row = 1

    # Title
    cell = ws.cell(row=row, column=1, value="QuantPricer — Pricing Report")
    cell.font = Font(name="Arial", bold=True, color=NAVY, size=16)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = \
        Font(name="Arial", color="888888", size=10, italic=True)
    row += 2

    # ── NPV Result ───────────────────────────────────────────
    row = _section_header(ws, row, "RESULT", 5)
    npv = result.get("npv", 0)
    row = _write_label_value(ws, row, "Net Present Value", npv,
                             val_font=NPV_FONT, val_fmt=DOLLAR_FMT)
    row = _write_label_value(ws, row, "Computation Time",
                             f"{result.get('elapsed_ms', 0)} ms")
    row = _write_label_value(ws, row, "Trade ID", result.get("trade_id", ""))
    row += 1

    # ── Instrument Details ───────────────────────────────────
    row = _section_header(ws, row, "INSTRUMENT", 5)
    inst_type = instrument.get("type", "")
    params = instrument.get("params", {})

    row = _write_label_value(ws, row, "Instrument Type", inst_type.upper(),
                             val_font=INPUT_FONT)

    # Write all instrument params
    for key, val in params.items():
        if key.startswith("_"):
            continue
        display_key = key.replace("_", " ").title()
        row = _write_label_value(ws, row, display_key, val, val_font=INPUT_FONT)
    row += 1

    # ── Market Data ──────────────────────────────────────────
    row = _section_header(ws, row, "MARKET DATA", 5)
    md = market_data or {}

    row = _write_label_value(ws, row, "Pricing Date",
                             md.get("pricing_date", ""), val_font=INPUT_FONT)

    spots = md.get("spot_prices", {})
    for sym, px in spots.items():
        row = _write_label_value(ws, row, f"Spot ({sym})", px,
                                 val_font=INPUT_FONT, val_fmt=NUMBER_FMT)

    vols = md.get("flat_vols", {})
    for sym, v in vols.items():
        row = _write_label_value(ws, row, f"Volatility ({sym})", v,
                                 val_font=INPUT_FONT, val_fmt=PCT_FMT)

    divs = md.get("dividend_yields", {})
    for sym, d in divs.items():
        row = _write_label_value(ws, row, f"Div Yield ({sym})", d,
                                 val_font=INPUT_FONT, val_fmt=PCT_FMT)

    rate_points = md.get("rate_curve", md.get("rate_curve_points", []))
    if isinstance(rate_points, list) and rate_points:
        for pt in rate_points[:5]:
            tenor = pt.get("tenor", "")
            r = pt.get("rate", 0)
            row = _write_label_value(ws, row, f"Rate ({tenor})", r,
                                     val_font=INPUT_FONT, val_fmt=PCT_FMT)
    row += 1

    # ── Model & Engine ───────────────────────────────────────
    row = _section_header(ws, row, "MODEL & ENGINE", 5)
    row = _write_label_value(ws, row, "Model", model, val_font=INPUT_FONT)
    row = _write_label_value(ws, row, "Engine", engine, val_font=INPUT_FONT)

    if engine_params:
        for k, v in engine_params.items():
            display = k.replace("_", " ").title()
            row = _write_label_value(ws, row, display, v, val_font=INPUT_FONT)
    row += 1

    # ── Greeks Summary (inline) ──────────────────────────────
    if greeks:
        row = _section_header(ws, row, "GREEKS SUMMARY", 5)
        greek_labels = {
            "delta": ("Delta", NUMBER_FMT),
            "gamma": ("Gamma", NUMBER_FMT),
            "vega": ("Vega", NUMBER_FMT),
            "theta": ("Theta", NUMBER_FMT),
            "rho": ("Rho", NUMBER_FMT),
            "dv01": ("DV01", NUMBER_FMT),
            "duration": ("Modified Duration", '0.0000'),
            "convexity": ("Convexity", '0.00'),
            "cs01": ("CS01", DOLLAR_FMT),
        }
        for key, (label, fmt) in greek_labels.items():
            val = greeks.get(key)
            if val is not None:
                row = _write_label_value(ws, row, label, val, val_fmt=fmt)


# ── Sheet 2: Greeks Detail ───────────────────────────────────

def _build_greeks_sheet(wb, greeks, result):
    ws = wb.create_sheet("Greeks Detail")
    ws.sheet_properties.tabColor = "2E86AB"
    _set_col_widths(ws, {"A": 20, "B": 18, "C": 18, "D": 18, "E": 18})

    row = 1
    row = _section_header(ws, row, "SENSITIVITY ANALYSIS", 5)
    row = _write_label_value(ws, row, "Base NPV", result.get("npv", 0),
                             val_fmt=DOLLAR_FMT)
    row += 1

    # Table header
    _write_header_row(ws, row, ["Greek", "Value", "Interpretation"])
    row += 1

    greek_info = {
        "delta": ("Delta", "dP per $1 move in underlying"),
        "gamma": ("Gamma", "dDelta per $1 move in underlying"),
        "vega": ("Vega", "dP per 1% vol move"),
        "theta": ("Theta", "dP per 1 day"),
        "rho": ("Rho", "dP per 1% rate move"),
        "dv01": ("DV01", "dP per 1bp rate shift"),
        "duration": ("Mod Duration", "% price change per 1% rate move"),
        "convexity": ("Convexity", "2nd order rate sensitivity"),
        "cs01": ("CS01", "dP per 1bp spread shift"),
    }

    for key, (label, desc) in greek_info.items():
        val = greeks.get(key)
        if val is None:
            continue
        fill = ALT_ROW_FILL if (row % 2 == 0) else None
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = LABEL_FONT
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = VALUE_FONT
        c2.number_format = NUMBER_FMT
        c3 = ws.cell(row=row, column=3, value=desc)
        c3.font = Font(name="Arial", color="888888", size=10, italic=True)
        if fill:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = fill
        row += 1


# ── Sheet 3: MC Simulation Paths ─────────────────────────────

def _build_mc_paths_sheet(wb, mc_data, instrument):
    ws = wb.create_sheet("MC Simulations")
    ws.sheet_properties.tabColor = "E67E22"

    paths = mc_data.get("paths")  # shape: (num_paths, num_steps+1)
    payoffs = mc_data.get("payoffs")  # shape: (num_paths,)
    stats = mc_data.get("statistics", {})
    num_paths = mc_data.get("num_paths", 0)
    num_steps = mc_data.get("num_steps", 0)

    row = 1
    row = _section_header(ws, row, "MONTE CARLO SIMULATION RESULTS", num_steps + 5)

    # Statistics
    row = _write_label_value(ws, row, "Number of Paths", num_paths, val_fmt=INT_FMT)
    row = _write_label_value(ws, row, "Number of Steps", num_steps, val_fmt=INT_FMT)
    row = _write_label_value(ws, row, "Mean Payoff", stats.get("mean_payoff", 0),
                             val_fmt=DOLLAR_FMT)
    row = _write_label_value(ws, row, "Std Dev", stats.get("std_payoff", 0),
                             val_fmt=DOLLAR_FMT)
    row = _write_label_value(ws, row, "Discounted Mean (NPV)", stats.get("npv", 0),
                             val_fmt=DOLLAR_FMT)
    row = _write_label_value(ws, row, "Std Error", stats.get("std_error", 0),
                             val_fmt=NUMBER_FMT)
    row = _write_label_value(ws, row, "95% CI Lower", stats.get("ci_lower", 0),
                             val_fmt=DOLLAR_FMT)
    row = _write_label_value(ws, row, "95% CI Upper", stats.get("ci_upper", 0),
                             val_fmt=DOLLAR_FMT)
    row += 1

    # Payoff distribution header
    row = _section_header(ws, row, "PATH-WISE PAYOFFS", 4)

    # Write path prices and payoffs
    if paths is not None:
        # Limit display to first 500 paths for performance
        max_display = min(len(paths), 500)
        display_paths = paths[:max_display]
        display_payoffs = payoffs[:max_display] if payoffs is not None else None

        # Header row: Path#, S(t0), S(t1), ..., S(T), Payoff
        headers = ["Path #"]
        for t in range(display_paths.shape[1] if hasattr(display_paths, 'shape') else 0):
            if t == 0:
                headers.append("S(t=0)")
            elif t == (display_paths.shape[1] - 1 if hasattr(display_paths, 'shape') else 0):
                headers.append("S(T)")
            else:
                headers.append(f"S(t={t})")
        headers.append("Payoff")

        _write_header_row(ws, row, headers)
        row += 1

        # Data rows
        for i in range(max_display):
            ws.cell(row=row, column=1, value=i + 1).font = VALUE_FONT

            path_row = display_paths[i] if hasattr(display_paths, '__getitem__') else []
            for j, val in enumerate(path_row):
                c = ws.cell(row=row, column=2 + j, value=float(val))
                c.number_format = NUMBER_FMT
                c.font = VALUE_FONT

            if display_payoffs is not None:
                payoff_col = 2 + len(path_row)
                c = ws.cell(row=row, column=payoff_col, value=float(display_payoffs[i]))
                c.number_format = DOLLAR_FMT
                c.font = VALUE_FONT

            if i % 2 == 1:
                for col in range(1, payoff_col + 1 if display_payoffs is not None else 2 + len(path_row)):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

            row += 1

        if len(paths) > max_display:
            ws.cell(row=row + 1, column=1,
                    value=f"... {len(paths) - max_display} more paths not shown").font = \
                Font(name="Arial", color="888888", italic=True)

    # Set column widths
    ws.column_dimensions["A"].width = 10
    for col_idx in range(2, 50):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


# ── Sheet 4: MC Random Numbers ───────────────────────────────

def _build_mc_randoms_sheet(wb, mc_data):
    ws = wb.create_sheet("MC Random Numbers")
    ws.sheet_properties.tabColor = "8E44AD"

    randoms = mc_data.get("random_numbers")  # shape: (num_paths, num_steps)
    if randoms is None:
        return

    row = 1
    row = _section_header(ws, row, "STANDARD NORMAL RANDOM DRAWS (Z)", 20)
    ws.cell(row=row, column=1,
            value="These are the Z ~ N(0,1) draws used to generate each path step.").font = \
        Font(name="Arial", color="888888", italic=True, size=10)
    row += 2

    max_display = min(len(randoms), 500)
    display = randoms[:max_display]

    num_steps = display.shape[1] if hasattr(display, 'shape') and len(display.shape) > 1 else 0

    headers = ["Path #"] + [f"Z(t={t+1})" for t in range(num_steps)]
    _write_header_row(ws, row, headers)
    row += 1

    for i in range(max_display):
        ws.cell(row=row, column=1, value=i + 1).font = VALUE_FONT
        for j in range(num_steps):
            c = ws.cell(row=row, column=2 + j, value=float(display[i][j]))
            c.number_format = '0.000000'
            c.font = VALUE_FONT
        if i % 2 == 1:
            for col in range(1, num_steps + 2):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL
        row += 1

    ws.column_dimensions["A"].width = 10
    for col_idx in range(2, num_steps + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
