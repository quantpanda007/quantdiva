"""
Bulk upload endpoint — accepts Excel with FX forward deals + market data.

Market data is per currency pair:
  "{CCY_PAIR} Forward"  sheet → forward rate curve
  "{CCY_PAIR} Discount" sheet → discount factor curve

Deals sheet has a "Valuate" column (Yes/No) to include/exclude deals.

Pricing: curve-based (if market data sheets present) or flat-rate fallback.
"""

from __future__ import annotations

import io
import math
import logging
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()
logger = logging.getLogger(__name__)
print("[BULK_UPLOAD] === V2: MULTI-CCY + VALUATE + LINEAR-DEFAULT ===")


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "valuate": "valuate",
    "transaction ref no": "transaction_ref",
    "transaction ref": "transaction_ref",
    "trans ref": "transaction_ref",
    "ref no": "transaction_ref",
    "ref": "transaction_ref",
    "client name": "client_name",
    "client": "client_name",
    "counterparty a": "cpty_a",
    "counterparty b": "cpty_b",
    "counterp arty a": "cpty_a",
    "counterp arty b": "cpty_b",
    "direction a": "buy_contract",
    "direction b": "sell_contract",
    "transaction date / trade date": "trade_date",
    "transaction date": "trade_date",
    "trade date": "trade_date",
    "effective date": "effective_date",
    "valuation / reporting date": "reporting_date",
    "valuation date": "reporting_date",
    "reporting date": "reporting_date",
    "maturity date": "maturity_date",
    "maturity da": "maturity_date",
    "maturity": "maturity_date",
    "type of contract": "contract_type",
    "type": "contract_type",
    "spot": "spot",
    "strike rate": "strike",
    "strike": "strike",
    "currency pair": "ccy_pair",
    "currency pa": "ccy_pair",
    "ccy pair": "ccy_pair",
    "notional currency 1": "notional_1",
    "notional ccy 1": "notional_1",
    "notional 1": "notional_1",
    "notional": "notional_1",
    "notional ccy 2": "notional_2",
    "notional currency 2": "notional_2",
    "buy /sell": "direction_1",
    "buy/sell": "direction_1",
    "buy /sell (ccy1)": "direction_1",
    "buy/sell (ccy1)": "direction_1",
    "buy /sell (ccy2)": "direction_2",
    "buy/sell (ccy2)": "direction_2",
    "domestic rate": "domestic_rate",
    "foreign rate": "foreign_rate",
    # Range forward delivery dates
    "delivery start date": "delivery_start_date",
    "delivery start": "delivery_start_date",
    "start date": "delivery_start_date",
    "maturity start date": "delivery_start_date",
    "delivery end date": "delivery_end_date",
    "delivery end": "delivery_end_date",
    "end date": "delivery_end_date",
    "maturity end date": "delivery_end_date",
}


def _normalize_header(h: str) -> str:
    return h.strip().lower().replace("\n", " ").replace("  ", " ")


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _add_business_days(d: date, days: int) -> date:
    current = d
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


# ---------------------------------------------------------------------------
# Market Data — per currency pair
# ---------------------------------------------------------------------------

class MarketCurves:
    """Forward + discount curves for one currency pair."""

    def __init__(self):
        self.forward_points: List[Tuple[int, float]] = []
        self.discount_points: List[Tuple[int, float]] = []
        self._fwd_spline = None
        self._disc_spline = None

    def has_curves(self) -> bool:
        return len(self.forward_points) > 0 and len(self.discount_points) > 0

    def _build_splines(self):
        if self._fwd_spline is None and len(self.forward_points) >= 3:
            try:
                from scipy.interpolate import CubicSpline
                d = dict(self.forward_points)
                xs = sorted(d.keys())
                ys = [d[x] for x in xs]
                if len(xs) >= 3:
                    self._fwd_spline = CubicSpline(xs, ys, extrapolate=True)
            except ImportError:
                pass
        if self._disc_spline is None and len(self.discount_points) >= 3:
            try:
                from scipy.interpolate import CubicSpline
                d = dict(self.discount_points)
                xs = sorted(d.keys())
                ys = [d[x] for x in xs]
                if len(xs) >= 3:
                    self._disc_spline = CubicSpline(xs, ys, extrapolate=True)
            except ImportError:
                pass

    def interpolate_forward(self, maturity_date: date, method: str = "linear") -> float:
        x = maturity_date.toordinal()
        if method == "cubic_spline":
            self._build_splines()
            if self._fwd_spline is not None:
                return float(self._fwd_spline(x))
        return _linear_interpolate(self.forward_points, x)

    def interpolate_df(self, maturity_date: date, method: str = "linear") -> float:
        x = maturity_date.toordinal()
        if method == "cubic_spline":
            self._build_splines()
            if self._disc_spline is not None:
                return float(self._disc_spline(x))
        return _linear_interpolate(self.discount_points, x)


def _linear_interpolate(points: List[Tuple[int, float]], x: int) -> float:
    if not points:
        raise ValueError("No curve points for interpolation")
    if x <= points[0][0]:
        return points[0][1]
    if x >= points[-1][0]:
        return points[-1][1]
    for i in range(len(points) - 1):
        x0, y0 = points[i]
        x1, y1 = points[i + 1]
        if x0 <= x <= x1:
            return y0 + (y1 - y0) * (x - x0) / (x1 - x0)
    return points[-1][1]


def _parse_curve_sheet(ws) -> List[Tuple[int, float]]:
    """Parse a 2-column curve sheet (Date, Value). Returns [(ordinal, value)]."""
    headers = [str(c.value or "").strip().lower() for c in ws[1]]

    date_col = val_col = None
    for idx, h in enumerate(headers):
        if h == "date":
            date_col = idx
        elif h in ("forward rate", "forwards mid", "discount factor", "discount", "discount rate"):
            val_col = idx
        elif "forward" in h and ("mid" in h or "rate" in h):
            val_col = idx

    if date_col is None or val_col is None:
        return []

    points = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[date_col] if date_col < len(row) else None
        num_val = row[val_col] if val_col < len(row) else None
        if date_val is None or num_val is None:
            continue
        d = _parse_date(date_val)
        if d is None:
            continue
        points.append((d.toordinal(), float(num_val)))

    points.sort(key=lambda x: x[0])
    return points


def parse_all_market_data(wb: openpyxl.Workbook) -> Dict[str, MarketCurves]:
    """
    Find all '{CCY_PAIR} Forward' and '{CCY_PAIR} Discount' sheets.
    Also supports legacy single 'Forward' + 'Discount' sheets (mapped to USDINR).
    Returns dict: { "USDINR": MarketCurves, "EURINR": MarketCurves, ... }
    """
    curves_map: Dict[str, MarketCurves] = {}
    sheet_names = wb.sheetnames

    # Detect per-pair sheets: "{PAIR} Forward" / "{PAIR} Discount"
    fwd_sheets = {}
    disc_sheets = {}
    for name in sheet_names:
        lower = name.lower().strip()
        if lower.endswith(" forward"):
            pair = name[: -len(" Forward")].strip().upper().replace("/", "")
            fwd_sheets[pair] = name
        elif lower.endswith(" discount"):
            pair = name[: -len(" Discount")].strip().upper().replace("/", "")
            disc_sheets[pair] = name

    # Build curves for each pair that has both forward and discount
    for pair in fwd_sheets:
        if pair in disc_sheets:
            mc = MarketCurves()
            mc.forward_points = _parse_curve_sheet(wb[fwd_sheets[pair]])
            mc.discount_points = _parse_curve_sheet(wb[disc_sheets[pair]])
            if mc.has_curves():
                curves_map[pair] = mc
                print(f"[CURVES] {pair}: {len(mc.forward_points)} fwd + {len(mc.discount_points)} disc points")

    # Legacy fallback: sheets named exactly "Forward" and "Discount" → USDINR
    if not curves_map:
        lower_map = {s.lower(): s for s in sheet_names}
        if "forward" in lower_map and "discount" in lower_map:
            mc = MarketCurves()
            mc.forward_points = _parse_curve_sheet(wb[lower_map["forward"]])
            mc.discount_points = _parse_curve_sheet(wb[lower_map["discount"]])
            if mc.has_curves():
                curves_map["USDINR"] = mc
                print(f"[CURVES] Legacy mode: USDINR {len(mc.forward_points)} fwd + {len(mc.discount_points)} disc")

    return curves_map


# ---------------------------------------------------------------------------
# Parse deals
# ---------------------------------------------------------------------------

def parse_upload(file_bytes: bytes) -> Tuple[List[Dict], List[Dict], Dict[str, MarketCurves]]:
    """Parse Excel. Returns (deals, errors, curves_map)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    curves_map = parse_all_market_data(wb)

    # Find "Deals" sheet, or first non-market-data sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower() == "deals":
            ws = wb[name]
            break
    if ws is None:
        skip = {"forward", "discount", "instructions"}
        # Also skip per-pair market data sheets
        for name in wb.sheetnames:
            lower = name.lower().strip()
            if lower in skip or lower.endswith(" forward") or lower.endswith(" discount"):
                continue
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    headers = [_normalize_header(str(c.value or "")) for c in ws[1]]

    col_map = {}
    direction_cols = []
    for idx, h in enumerate(headers):
        matched = COLUMN_ALIASES.get(h)
        if matched:
            if matched == "direction_1":
                direction_cols.append(idx)
            else:
                col_map[matched] = idx
    if len(direction_cols) >= 1:
        col_map["direction_1"] = direction_cols[0]

    deals, errors = [], []
    has_curves = len(curves_map) > 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(field, _col_map=col_map, _row=row):
            idx = _col_map.get(field)
            if idx is not None and idx < len(_row):
                return _row[idx]
            return None

        # Valuate check — only process rows explicitly marked Yes
        valuate = str(get("valuate") or "").strip().lower()
        if valuate not in ("yes", "y", ""):
            continue

        deal = {
            "row": row_idx,
            "transaction_ref": str(get("transaction_ref") or f"ROW-{row_idx}"),
            "client_name": str(get("client_name") or ""),
            "cpty_a": str(get("cpty_a") or ""),
            "cpty_b": str(get("cpty_b") or ""),
            "buy_contract": str(get("buy_contract") or ""),
            "sell_contract": str(get("sell_contract") or ""),
            "trade_date": _parse_date(get("trade_date")),
            "effective_date": _parse_date(get("effective_date")),
            "reporting_date": _parse_date(get("reporting_date")),
            "contract_type": str(get("contract_type") or "Forward"),
            "spot": _parse_number(get("spot")),
            "strike": _parse_number(get("strike")),
            "ccy_pair": str(get("ccy_pair") or "USDINR").replace("/", "").upper(),
            "notional_1": _parse_number(get("notional_1")),
            "direction_1": str(get("direction_1") or "Sell"),
            "notional_2": _parse_number(get("notional_2")),
            "maturity_date": _parse_date(get("maturity_date")),
            "delivery_start_date": _parse_date(get("delivery_start_date")),
            "delivery_end_date": _parse_date(get("delivery_end_date")),
            "domestic_rate": _parse_number(get("domestic_rate")) if get("domestic_rate") else None,
            "foreign_rate": _parse_number(get("foreign_rate")) if get("foreign_rate") else None,
        }

        # Range forward: auto-compute maturity
        if deal["delivery_start_date"] and deal["delivery_end_date"]:
            direction = deal["direction_1"].lower()
            if direction in ("sell", "s"):
                deal["maturity_date"] = deal["delivery_start_date"]
            else:
                deal["maturity_date"] = deal["delivery_end_date"]

        errs = []
        if not deal["strike"]:
            errs.append("Missing strike rate")
        if not deal["notional_1"]:
            errs.append("Missing notional")
        if not deal["maturity_date"]:
            errs.append("Missing or invalid maturity date")
        if not deal["spot"] and not has_curves:
            errs.append("Missing spot rate (required without market data curves)")
        if has_curves and deal["ccy_pair"] not in curves_map:
            errs.append(f"No market data sheets for {deal['ccy_pair']} "
                        f"(need '{deal['ccy_pair']} Forward' and '{deal['ccy_pair']} Discount')")

        if errs:
            errors.append({"row": row_idx, "ref": deal["transaction_ref"], "errors": "; ".join(errs)})
        else:
            deals.append(deal)

    return deals, errors, curves_map


# ---------------------------------------------------------------------------
# Pricing — Curve-based (per currency pair)
# ---------------------------------------------------------------------------

def price_deals_with_curves(
    deals: List[Dict],
    curves_map: Dict[str, MarketCurves],
    interpolation: str = "linear",
) -> Tuple[List[Dict], List[Dict]]:
    """
    NPV = Notional × (Strike - Forward) × DF × sign
    Curves looked up per deal's currency pair.
    """
    results, pricing_errors = [], []

    for deal in deals:
        try:
            pricing_date = deal["reporting_date"] or date.today()

            if deal["maturity_date"] <= pricing_date:
                results.append({
                    **deal, "npv": 0.0, "forward_rate": deal["spot"],
                    "discount_factor": 1.0, "long_term": 0.0, "short_term": 0.0,
                    "pricing_method": "curve", "pricing_date": pricing_date, "status": "Expired",
                })
                continue

            # Get curves for this deal's currency pair
            ccy = deal["ccy_pair"]
            curves = curves_map.get(ccy)
            if not curves:
                pricing_errors.append({
                    "row": deal["row"], "ref": deal["transaction_ref"],
                    "errors": f"No curves for {ccy}",
                })
                continue

            maturity = deal["maturity_date"]
            days_to_maturity = (maturity - pricing_date).days

            fwd_rate = curves.interpolate_forward(maturity, method=interpolation)
            disc_factor = curves.interpolate_df(maturity, method=interpolation)

            # Precise strike from CCY2/CCY1
            strike = deal["strike"]
            if deal["notional_2"] and deal["notional_1"]:
                strike = deal["notional_2"] / deal["notional_1"]

            direction = deal["direction_1"].lower()
            sign = 1.0 if direction in ("sell", "s") else -1.0

            npv = deal["notional_1"] * (strike - fwd_rate) * disc_factor * sign

            months = days_to_maturity / 30.0
            results.append({
                **deal, "npv": npv, "strike_precise": strike, "forward_rate": fwd_rate,
                "discount_factor": disc_factor,
                "long_term": npv if months > 12 else 0.0,
                "short_term": npv if months <= 12 else 0.0,
                "pricing_method": "curve", "pricing_date": pricing_date, "status": "OK",
            })

        except Exception as e:
            pricing_errors.append({"row": deal["row"], "ref": deal["transaction_ref"], "errors": str(e)})

    return results, pricing_errors


# ---------------------------------------------------------------------------
# Pricing — Flat rate fallback
# ---------------------------------------------------------------------------

def price_deals_flat(
    deals: List[Dict],
    default_domestic_rate: float,
    default_foreign_rate: float,
) -> Tuple[List[Dict], List[Dict]]:
    """Flat-rate pricing via PricingService. Fallback when no curves provided."""
    from services.pricers.pricing_service import PricingService
    from api.v1.helpers import build_instrument_from_request, build_market_env_from_request
    from api.v1.schemas import InstrumentRequest, MarketDataRequest, UnderlyingData

    ps = PricingService()
    results, pricing_errors = [], []

    for deal in deals:
        try:
            r_d = deal["domestic_rate"] if deal["domestic_rate"] is not None else default_domestic_rate
            r_f = deal["foreign_rate"] if deal["foreign_rate"] is not None else default_foreign_rate
            pricing_date = deal["reporting_date"] or date.today()
            direction = "sell" if deal["direction_1"].lower() in ("sell", "s") else "buy"

            inst_req = InstrumentRequest(
                type="fx_forward",
                params={
                    "ccy_pair": deal["ccy_pair"], "strike": deal["strike"],
                    "delivery_date": deal["maturity_date"].isoformat(),
                    "notional": deal["notional_1"], "direction": direction,
                }
            )
            md_req = MarketDataRequest(
                pricing_date=pricing_date.isoformat(),
                underlyings={deal["ccy_pair"]: UnderlyingData(spot=deal["spot"], vol=0.06)},
                rate_curve=[{"tenor": "1Y", "rate": r_d}],
                foreign_rate=r_f,
            )

            instrument = build_instrument_from_request(inst_req)
            market_env = build_market_env_from_request(md_req, underlying=getattr(instrument, "ccy_pair", ""))
            result = ps.price(instrument, market_env, model_type="black_scholes", engine_type="analytic")

            T = (deal["maturity_date"] - pricing_date).days / 365.0
            fwd_rate = deal["spot"] * math.exp((r_d - r_f) * T) if T > 0 else deal["spot"]
            disc_factor = math.exp(-r_d * T) if T > 0 else 1.0
            months = (deal["maturity_date"] - pricing_date).days / 30.0

            results.append({
                **deal, "npv": result.npv, "forward_rate": fwd_rate,
                "discount_factor": disc_factor,
                "long_term": result.npv if months > 12 else 0.0,
                "short_term": result.npv if months <= 12 else 0.0,
                "domestic_rate_used": r_d, "foreign_rate_used": r_f,
                "pricing_method": "flat", "pricing_date": pricing_date, "status": "OK",
            })

        except Exception as e:
            pricing_errors.append({"row": deal["row"], "ref": deal["transaction_ref"], "errors": str(e)})

    return results, pricing_errors


# ---------------------------------------------------------------------------
# Generate results Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1a2540")
HEADER_FONT = Font(bold=True, color="F59E0B", size=11)
NUM_FMT_MONEY = '#,##0.00'
NUM_FMT_RATE = '0.0000'
NUM_FMT_DF = '0.000000'
THIN_BORDER = Border(bottom=Side(style="thin", color="2E4A7A"))


def generate_results_excel(results, errors, upload_filename, pricing_method="flat"):
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35

    active = [r for r in results if r.get("status") != "Expired"]
    expired = len(results) - len(active)
    spot = (active[0]["spot"] if active else 85.47) or 85.47

    # Currency pairs in this upload
    ccy_pairs = sorted(set(r.get("ccy_pair", "USDINR") for r in active))

    rows = [
        ("Upload File", upload_filename),
        ("Pricing Method", "Curve-based" if pricing_method == "curve" else "Flat Rate"),
        ("Currency Pairs", ", ".join(ccy_pairs)),
        ("Total Deals", len(results) + len(errors)),
        ("Priced OK", len(active)),
        ("Expired", expired),
        ("Errors", len(errors)),
    ]
    if results:
        rows.append(("Pricing Date", str(results[0].get("pricing_date", ""))))
    rows += [
        ("", ""),
        ("Total NPV (INR)", sum(r["npv"] for r in active)),
        ("Total NPV (USD)", sum(r["npv"] for r in active) / spot if spot else 0),
        ("Total Long Term", sum(r["long_term"] for r in active)),
        ("Total Short Term", sum(r["short_term"] for r in active)),
        ("", ""),
        ("Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for i, (label, val) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        cell = ws.cell(row=i, column=2, value=val)
        if isinstance(val, float):
            cell.number_format = NUM_FMT_MONEY

    # --- Deal Results ---
    ws2 = wb.create_sheet("Deal_Results")
    headers = [
        "Transaction Ref", "Client Name", "Cpty A", "Cpty B",
        "Ccy Pair", "Strike Rate", "Notional (CCY1)", "Direction",
        "Trade Date", "Effective Date", "Maturity Date", "Pricing Date",
        "Spot", "Forward Rate", "Discount Factor",
        "NPV (INR)", "NPV (USD)", "Long Term", "Short Term", "Status",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for ri, r in enumerate(results, 2):
        s = r.get("spot", 85.47) or 85.47
        data = [
            r["transaction_ref"], r["client_name"], r["cpty_a"], r["cpty_b"],
            r["ccy_pair"], r.get("strike_precise", r["strike"]), r["notional_1"], r["direction_1"],
            r["trade_date"].isoformat() if r.get("trade_date") else "",
            r["effective_date"].isoformat() if r.get("effective_date") else "",
            r["maturity_date"].isoformat() if r.get("maturity_date") else "",
            str(r.get("pricing_date", "")),
            r["spot"], r["forward_rate"], r["discount_factor"],
            r["npv"], r["npv"] / s if s else 0, r["long_term"], r["short_term"], r["status"],
        ]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=ri, column=col, value=val)
            cell.border = THIN_BORDER
            if col in (6, 13, 14):
                cell.number_format = NUM_FMT_RATE
            elif col == 7:
                cell.number_format = '#,##0'
            elif col == 15:
                cell.number_format = NUM_FMT_DF
            elif col in (16, 17, 18, 19):
                cell.number_format = NUM_FMT_MONEY
                if isinstance(val, (int, float)):
                    if r.get("status") == "Expired":
                        cell.font = Font(color="94A3B8")
                    elif val >= 0:
                        cell.font = Font(color="10B981")
                    else:
                        cell.font = Font(color="EF4444")
            elif col == 20:
                cell.font = Font(color="94A3B8") if r.get("status") == "Expired" else Font(color="10B981")

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(14, len(headers[col-1]) + 4)

    # --- Errors ---
    if errors:
        ws3 = wb.create_sheet("Errors")
        for col, h in enumerate(["Row", "Transaction Ref", "Error"], 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for i, e in enumerate(errors, 2):
            ws3.cell(row=i, column=1, value=e["row"])
            ws3.cell(row=i, column=2, value=e["ref"])
            ws3.cell(row=i, column=3, value=e["errors"]).font = Font(color="EF4444")
        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 25
        ws3.column_dimensions["C"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/bulk-upload")
async def bulk_upload(
    file: UploadFile = File(...),
    domestic_rate: float = Query(0.065, description="Fallback domestic rate"),
    foreign_rate: float = Query(0.045, description="Fallback foreign rate"),
    interpolation: str = Query("linear", description="Interpolation method: 'linear' or 'cubic_spline'"),
):
    """
    Upload Excel with FX forward deals + market data.
    Market data sheets: '{CCY_PAIR} Forward' and '{CCY_PAIR} Discount'.
    Falls back to flat rates if no market data sheets found.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files accepted")

    if interpolation not in ("linear", "cubic_spline"):
        raise HTTPException(400, "interpolation must be 'linear' or 'cubic_spline'")

    try:
        contents = await file.read()
        deals, parse_errors, curves_map = parse_upload(contents)

        if not deals and parse_errors:
            raise HTTPException(400, f"No valid deals found. Errors: {parse_errors}")

        if curves_map:
            print(f"[PRICING] Curve-based ({interpolation}), pairs: {list(curves_map.keys())}")
            results, pricing_errors = price_deals_with_curves(deals, curves_map, interpolation)
            method = "curve"
        else:
            print("[PRICING] Flat-rate (no curves)")
            results, pricing_errors = price_deals_flat(deals, domestic_rate, foreign_rate)
            method = "flat"

        result_bytes = generate_results_excel(results, parse_errors + pricing_errors, file.filename, method)

        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Optima_BulkResults_{file.filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Bulk upload failed")
        raise HTTPException(500, f"Bulk upload failed: {e}")


@router.get("/bulk-template")
async def download_template():
    """Download blank template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"
    ws.cell(row=1, column=1, value="Download templates from Optima UI")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Optima_BulkUpload_Template.xlsx"'},
    )
