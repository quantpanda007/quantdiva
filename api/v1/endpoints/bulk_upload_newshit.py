"""
Bulk upload endpoint — accepts Excel file with FX forward deals,
prices each row using either:
  1. Curve-based pricing (if Forward + Discount sheets present)
  2. Flat-rate pricing (fallback to domestic/foreign rate inputs)

Returns results as downloadable Excel.
"""

from __future__ import annotations

import io
import math
import logging
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column mapping — maps business team's Excel headers to internal field names
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "transaction ref no": "transaction_ref",
    "transaction ref": "transaction_ref",
    "trans ref": "transaction_ref",
    "ref no": "transaction_ref",
    "ref": "transaction_ref",
    "client name": "client_name",
    "client": "client_name",
    "counterparty a": "cpty_a",
    "counterparty b": "cpty_b",
    "counterp arty a": "cpty_a",
    "counterp arty b": "cpty_b",
    "buy contract": "buy_contract",
    "sell contract": "sell_contract",
    "transaction date / trade date": "trade_date",
    "transaction date": "trade_date",
    "trade date": "trade_date",
    "effective date": "effective_date",
    "valuation / reporting date": "reporting_date",
    "valuation date": "reporting_date",
    "reporting date": "reporting_date",
    "maturity date": "maturity_date",
    "maturity da": "maturity_date",
    "maturity": "maturity_date",
    "type of contract": "contract_type",
    "type": "contract_type",
    "spot": "spot",
    "strike rate": "strike",
    "strike": "strike",
    "currency pair": "ccy_pair",
    "currency pa": "ccy_pair",
    "ccy pair": "ccy_pair",
    "notional currency 1": "notional_1",
    "notional ccy 1": "notional_1",
    "notional 1": "notional_1",
    "notional": "notional_1",
    "notional ccy 2": "notional_2",
    "notional currency 2": "notional_2",
    "buy /sell": "direction_1",
    "buy/sell": "direction_1",
    "domestic rate": "domestic_rate",
    "foreign rate": "foreign_rate",
    # Delivery window (range forward)
    "delivery start date": "delivery_start_date",
    "delivery start": "delivery_start_date",
    "start date": "delivery_start_date",
    "delivery end date": "delivery_end_date",
    "delivery end": "delivery_end_date",
    "end date": "delivery_end_date",
}


def _normalize_header(h: str) -> str:
    return h.strip().lower().replace("\n", " ").replace("  ", " ")


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # Try US format (M/D/Y) before European (D/M/Y) since Excel market data uses US dates
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _add_business_days(d: date, days: int) -> date:
    current = d
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


# ---------------------------------------------------------------------------
# Market Data Curve Parser
# ---------------------------------------------------------------------------

class MarketCurves:
    """
    Holds forward and discount curves parsed from Excel sheets.
    
    Stores points as (date_ordinal, value) for date-based interpolation
    matching Excel's FORECAST.LINEAR behavior.
    """

    def __init__(self):
        self.forward_points: List[Tuple[int, float]] = []  # (date_ordinal, fwd_rate)
        self.forward_dates: List[date] = []                  # actual dates for reference
        self.discount_points: List[Tuple[int, float]] = []   # (date_ordinal, df)
        self.discount_dates: List[date] = []
        self.spot_rate: float = 0.0
        self._fwd_spline = None
        self._disc_spline = None

    def has_curves(self) -> bool:
        return len(self.forward_points) > 0 and len(self.discount_points) > 0

    def _build_splines(self):
        """Build cubic spline objects (lazy, on first use)."""
        if self._fwd_spline is None and len(self.forward_points) >= 3:
            try:
                from scipy.interpolate import CubicSpline
                fwd_dedup = dict(self.forward_points)
                xs = sorted(fwd_dedup.keys())
                ys = [fwd_dedup[x] for x in xs]
                if len(xs) >= 3:
                    self._fwd_spline = CubicSpline(xs, ys, extrapolate=True)
            except ImportError:
                logger.warning("scipy not available, cubic spline disabled")
        if self._disc_spline is None and len(self.discount_points) >= 3:
            try:
                from scipy.interpolate import CubicSpline
                disc_dedup = dict(self.discount_points)
                xs = sorted(disc_dedup.keys())
                ys = [disc_dedup[x] for x in xs]
                if len(xs) >= 3:
                    self._disc_spline = CubicSpline(xs, ys, extrapolate=True)
            except ImportError:
                pass

    def interpolate_forward(self, maturity_date: date, method: str = "linear") -> float:
        """Interpolate forward rate at maturity date."""
        x = maturity_date.toordinal()
        if method == "cubic_spline":
            self._build_splines()
            if self._fwd_spline is not None:
                return float(self._fwd_spline(x))
        return _linear_interpolate(self.forward_points, x)

    def interpolate_df(self, maturity_date: date, method: str = "linear") -> float:
        """Interpolate discount factor at maturity date."""
        x = maturity_date.toordinal()
        if method == "cubic_spline":
            self._build_splines()
            if self._disc_spline is not None:
                return float(self._disc_spline(x))
        return _linear_interpolate(self.discount_points, x)


def _linear_interpolate(points: List[Tuple[int, float]], x: int) -> float:
    if not points:
        raise ValueError("No curve points for interpolation")
    if x <= points[0][0]:
        return points[0][1]
    if x >= points[-1][0]:
        return points[-1][1]
    for i in range(len(points) - 1):
        x0, y0 = points[i]
        x1, y1 = points[i + 1]
        if x0 <= x <= x1:
            return y0 + (y1 - y0) * (x - x0) / (x1 - x0)
    return points[-1][1]


def _parse_term_to_days(term: str) -> Optional[int]:
    term = term.strip().upper()
    parts = term.split()
    if len(parts) != 2:
        return None
    try:
        num = int(float(parts[0]))
    except ValueError:
        return None
    unit = parts[1]
    if unit in ("DY", "D", "DAYS", "DAY"):
        return num
    elif unit in ("YR", "Y", "YEAR", "YEARS"):
        return num * 365
    elif unit in ("MO", "M", "MONTH", "MONTHS"):
        return num * 30
    return None


def parse_market_data(wb: openpyxl.Workbook) -> Optional[MarketCurves]:
    """Parse Forward and Discount sheets. Returns MarketCurves or None."""
    sheet_map = {s.lower(): s for s in wb.sheetnames}

    if "forward" not in sheet_map or "discount" not in sheet_map:
        return None

    curves = MarketCurves()

    # --- Forward sheet ---
    ws_fwd = wb[sheet_map["forward"]]
    headers = [str(c.value or "").strip().lower() for c in ws_fwd[1]]

    date_col = tenor_col = fwd_col = days_col = None
    for idx, h in enumerate(headers):
        if h == "date":
            date_col = idx
        elif h == "days":
            days_col = idx
        elif "forward" in h and "mid" in h:
            fwd_col = idx
        elif h == "t":
            tenor_col = idx

    if fwd_col is None:
        logger.warning("Forward sheet: missing 'Forwards Mid' column")
        return None
    if date_col is None and days_col is None:
        logger.warning("Forward sheet: missing both 'Date' and 'Days' columns")
        return None

    for row in ws_fwd.iter_rows(min_row=2, values_only=True):
        fwd_val = row[fwd_col] if fwd_col < len(row) else None
        tenor_val = row[tenor_col] if tenor_col is not None and tenor_col < len(row) else None

        if fwd_val is None:
            continue

        fwd = float(fwd_val)

        if tenor_val and str(tenor_val).strip().upper() == "SP":
            curves.spot_rate = fwd

        # Get date — prefer Date column, fall back to Days column
        fwd_date = None
        if date_col is not None and date_col < len(row) and row[date_col] is not None:
            fwd_date = _parse_date(row[date_col])

        if fwd_date is not None:
            curves.forward_points.append((fwd_date.toordinal(), fwd))
            curves.forward_dates.append(fwd_date)
        elif days_col is not None and days_col < len(row) and row[days_col] is not None:
            days = int(float(row[days_col]))
            curves.forward_points.append((days, fwd))

    curves.forward_points.sort(key=lambda x: x[0])

    # --- Discount sheet ---
    ws_disc = wb[sheet_map["discount"]]
    headers = [str(c.value or "").strip().lower() for c in ws_disc[1]]

    date_col_d = term_col = df_col = None
    for idx, h in enumerate(headers):
        if h == "date":
            date_col_d = idx
        elif h in ("tenor", "term"):
            term_col = idx
        elif h in ("discount", "discount rate"):
            df_col = idx

    if df_col is None:
        logger.warning("Discount sheet: missing 'Discount' column")
        return None
    if date_col_d is None and term_col is None:
        logger.warning("Discount sheet: missing both 'Date' and 'Term' columns")
        return None

    for row in ws_disc.iter_rows(min_row=2, values_only=True):
        df_val = row[df_col] if df_col < len(row) else None
        if df_val is None:
            continue

        # Get date — prefer Date column, fall back to Term column
        disc_date = None
        if date_col_d is not None and date_col_d < len(row) and row[date_col_d] is not None:
            disc_date = _parse_date(row[date_col_d])

        if disc_date is None and term_col is not None and term_col < len(row) and row[term_col] is not None:
            term_val = row[term_col]
            disc_date = _parse_date(term_val)
            if disc_date is None:
                days = _parse_term_to_days(str(term_val))
                if days is not None:
                    curves.discount_points.append((days, float(df_val)))
                    continue

        if disc_date is None:
            continue

        curves.discount_points.append((disc_date.toordinal(), float(df_val)))
        curves.discount_dates.append(disc_date)

    curves.discount_points.sort(key=lambda x: x[0])

    if curves.has_curves():
        logger.info(
            f"Market curves: {len(curves.forward_points)} fwd points, "
            f"{len(curves.discount_points)} disc points, spot={curves.spot_rate}"
        )
        return curves

    return None


# ---------------------------------------------------------------------------
# Parse deals
# ---------------------------------------------------------------------------

def parse_upload(file_bytes: bytes) -> Tuple[List[Dict], List[Dict], Optional[MarketCurves]]:
    """Parse Excel. Returns (deals, errors, market_curves)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    market_curves = parse_market_data(wb)

    # Find deals sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower() not in ("forward", "discount"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    headers = [_normalize_header(str(c.value or "")) for c in ws[1]]

    col_map = {}
    direction_cols = []
    for idx, h in enumerate(headers):
        matched = COLUMN_ALIASES.get(h)
        if matched:
            if matched == "direction_1":
                direction_cols.append(idx)
            else:
                col_map[matched] = idx
    if len(direction_cols) >= 1:
        col_map["direction_1"] = direction_cols[0]

    deals, errors = [], []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        deal = {
            "row": row_idx,
            "transaction_ref": str(get("transaction_ref") or f"ROW-{row_idx}"),
            "client_name": str(get("client_name") or ""),
            "cpty_a": str(get("cpty_a") or ""),
            "cpty_b": str(get("cpty_b") or ""),
            "buy_contract": str(get("buy_contract") or ""),
            "sell_contract": str(get("sell_contract") or ""),
            "trade_date": _parse_date(get("trade_date")),
            "effective_date": _parse_date(get("effective_date")),
            "reporting_date": _parse_date(get("reporting_date")),
            "contract_type": str(get("contract_type") or "Forward"),
            "spot": _parse_number(get("spot")),
            "strike": _parse_number(get("strike")),
            "ccy_pair": str(get("ccy_pair") or "USDINR").replace("/", ""),
            "notional_1": _parse_number(get("notional_1")),
            "direction_1": str(get("direction_1") or "Sell"),
            "notional_2": _parse_number(get("notional_2")),
            "maturity_date": _parse_date(get("maturity_date")),
            "delivery_start_date": _parse_date(get("delivery_start_date")),
            "delivery_end_date": _parse_date(get("delivery_end_date")),
            "domestic_rate": _parse_number(get("domestic_rate")) if get("domestic_rate") else None,
            "foreign_rate": _parse_number(get("foreign_rate")) if get("foreign_rate") else None,
        }

        # Range forward: auto-compute maturity from delivery window + direction
        # Sell → start date (worst for seller), Buy → end date (worst for buyer)
        if deal["delivery_start_date"] and deal["delivery_end_date"]:
            direction = deal["direction_1"].lower()
            if direction in ("sell", "s"):
                deal["maturity_date"] = deal["delivery_start_date"]
            else:
                deal["maturity_date"] = deal["delivery_end_date"]

        errs = []
        if not deal["strike"]:
            errs.append("Missing strike rate")
        if not deal["notional_1"]:
            errs.append("Missing notional")
        if not deal["maturity_date"]:
            errs.append("Missing or invalid maturity date")
        if not deal["spot"]:
            errs.append("Missing spot rate")

        if errs:
            errors.append({"row": row_idx, "ref": deal["transaction_ref"], "errors": "; ".join(errs)})
        else:
            deals.append(deal)

    return deals, errors, market_curves


# ---------------------------------------------------------------------------
# Pricing — Curve-based
# ---------------------------------------------------------------------------

def price_deals_with_curves(
    deals: List[Dict],
    curves: MarketCurves,
    interpolation: str = "linear",
) -> Tuple[List[Dict], List[Dict]]:
    """
    NPV = Notional × (Strike - Forward) × DF × sign
    Forward and DF interpolated from curves at maturity.
    interpolation: 'linear' or 'cubic_spline'
    """
    results, pricing_errors = [], []

    for deal in deals:
        try:
            pricing_date = deal["reporting_date"] or date.today()
            spot_date = _add_business_days(pricing_date, 2)

            if deal["maturity_date"] <= pricing_date:
                results.append({
                    **deal, "npv": 0.0, "forward_rate": deal["spot"],
                    "discount_factor": 1.0, "long_term": 0.0, "short_term": 0.0,
                    "pricing_method": "curve", "pricing_date": pricing_date, "status": "Expired",
                })
                continue

            maturity = deal["maturity_date"]
            days_to_maturity = (maturity - pricing_date).days

            fwd_rate = curves.interpolate_forward(maturity, method=interpolation)
            disc_factor = curves.interpolate_df(maturity, method=interpolation)

            # Precise strike from CCY2/CCY1 notionals (avoids rounded Strike Rate column)
            strike = deal["strike"]
            if deal["notional_2"] and deal["notional_1"]:
                strike = deal["notional_2"] / deal["notional_1"]

            direction = deal["direction_1"].lower()
            sign = 1.0 if direction in ("sell", "s") else -1.0

            npv = deal["notional_1"] * (strike - fwd_rate) * disc_factor * sign

            months = days_to_maturity / 30.0
            results.append({
                **deal, "npv": npv, "strike_precise": strike, "forward_rate": fwd_rate,
                "discount_factor": disc_factor,
                "long_term": npv if months > 12 else 0.0,
                "short_term": npv if months <= 12 else 0.0,
                "pricing_method": "curve", "pricing_date": pricing_date, "status": "OK",
            })

        except Exception as e:
            pricing_errors.append({"row": deal["row"], "ref": deal["transaction_ref"], "errors": str(e)})

    return results, pricing_errors


# ---------------------------------------------------------------------------
# Pricing — Flat rate fallback
# ---------------------------------------------------------------------------

def price_deals_flat(
    deals: List[Dict],
    default_domestic_rate: float,
    default_foreign_rate: float,
) -> Tuple[List[Dict], List[Dict]]:
    """Flat-rate pricing via PricingService. Fallback when no curves provided."""
    from services.pricers.pricing_service import PricingService
    from api.v1.helpers import build_instrument_from_request, build_market_env_from_request
    from api.v1.schemas import InstrumentRequest, MarketDataRequest, UnderlyingData

    ps = PricingService()
    results, pricing_errors = [], []

    for deal in deals:
        try:
            r_d = deal["domestic_rate"] if deal["domestic_rate"] is not None else default_domestic_rate
            r_f = deal["foreign_rate"] if deal["foreign_rate"] is not None else default_foreign_rate
            pricing_date = deal["reporting_date"] or date.today()
            direction = "sell" if deal["direction_1"].lower() in ("sell", "s") else "buy"

            inst_req = InstrumentRequest(
                type="fx_forward",
                params={
                    "ccy_pair": deal["ccy_pair"], "strike": deal["strike"],
                    "delivery_date": deal["maturity_date"].isoformat(),
                    "notional": deal["notional_1"], "direction": direction,
                }
            )
            md_req = MarketDataRequest(
                pricing_date=pricing_date.isoformat(),
                underlyings={deal["ccy_pair"]: UnderlyingData(spot=deal["spot"], vol=0.06)},
                rate_curve=[{"tenor": "1Y", "rate": r_d}],
                foreign_rate=r_f,
            )

            instrument = build_instrument_from_request(inst_req)
            market_env = build_market_env_from_request(md_req, underlying=getattr(instrument, "ccy_pair", ""))
            result = ps.price(instrument, market_env, model_type="black_scholes", engine_type="analytic")

            T = (deal["maturity_date"] - pricing_date).days / 365.0
            fwd_rate = deal["spot"] * math.exp((r_d - r_f) * T) if T > 0 else deal["spot"]
            disc_factor = math.exp(-r_d * T) if T > 0 else 1.0
            months = (deal["maturity_date"] - pricing_date).days / 30.0

            results.append({
                **deal, "npv": result.npv, "forward_rate": fwd_rate,
                "discount_factor": disc_factor,
                "long_term": result.npv if months > 12 else 0.0,
                "short_term": result.npv if months <= 12 else 0.0,
                "domestic_rate_used": r_d, "foreign_rate_used": r_f,
                "pricing_method": "flat", "pricing_date": pricing_date, "status": "OK",
            })

        except Exception as e:
            pricing_errors.append({"row": deal["row"], "ref": deal["transaction_ref"], "errors": str(e)})

    return results, pricing_errors


# ---------------------------------------------------------------------------
# Generate results Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1a2540")
HEADER_FONT = Font(bold=True, color="F59E0B", size=11)
NUM_FMT_MONEY = '#,##0.00'
NUM_FMT_RATE = '0.0000'
NUM_FMT_DF = '0.000000'
THIN_BORDER = Border(bottom=Side(style="thin", color="2E4A7A"))


def generate_results_excel(results, errors, upload_filename, pricing_method="flat"):
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35

    active = [r for r in results if r.get("status") != "Expired"]
    expired = len(results) - len(active)
    spot = active[0]["spot"] if active else 85.47

    rows = [
        ("Upload File", upload_filename),
        ("Pricing Method", "Curve-based" if pricing_method == "curve" else "Flat Rate"),
        ("Total Deals", len(results) + len(errors)),
        ("Priced OK", len(active)),
        ("Expired", expired),
        ("Errors", len(errors)),
    ]
    if results:
        rows.append(("Pricing Date", str(results[0].get("pricing_date", ""))))
    rows += [
        ("", ""),
        ("Total NPV (INR)", sum(r["npv"] for r in active)),
        ("Total NPV (USD)", sum(r["npv"] for r in active) / spot),
        ("Total Long Term", sum(r["long_term"] for r in active)),
        ("Total Short Term", sum(r["short_term"] for r in active)),
        ("", ""),
        ("Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for i, (label, val) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        cell = ws.cell(row=i, column=2, value=val)
        if isinstance(val, float):
            cell.number_format = NUM_FMT_MONEY

    # --- Deal Results ---
    ws2 = wb.create_sheet("Deal_Results")
    headers = [
        "Transaction Ref", "Client Name", "Cpty A", "Cpty B",
        "Ccy Pair", "Strike Rate", "Notional (CCY1)", "Direction",
        "Trade Date", "Effective Date", "Maturity Date", "Pricing Date",
        "Spot", "Forward Rate", "Discount Factor",
        "NPV (INR)", "NPV (USD)", "Long Term", "Short Term", "Status",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for ri, r in enumerate(results, 2):
        s = r.get("spot", 85.47) or 85.47
        data = [
            r["transaction_ref"], r["client_name"], r["cpty_a"], r["cpty_b"],
            r["ccy_pair"], r.get("strike_precise", r["strike"]), r["notional_1"], r["direction_1"],
            r["trade_date"].isoformat() if r.get("trade_date") else "",
            r["effective_date"].isoformat() if r.get("effective_date") else "",
            r["maturity_date"].isoformat() if r.get("maturity_date") else "",
            str(r.get("pricing_date", "")),
            r["spot"], r["forward_rate"], r["discount_factor"],
            r["npv"], r["npv"] / s, r["long_term"], r["short_term"], r["status"],
        ]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=ri, column=col, value=val)
            cell.border = THIN_BORDER
            if col in (6, 13, 14):
                cell.number_format = NUM_FMT_RATE
            elif col == 7:
                cell.number_format = '#,##0'
            elif col == 15:
                cell.number_format = NUM_FMT_DF
            elif col in (16, 17, 18, 19):
                cell.number_format = NUM_FMT_MONEY
                if isinstance(val, (int, float)):
                    if r.get("status") == "Expired":
                        cell.font = Font(color="94A3B8")
                    elif val >= 0:
                        cell.font = Font(color="10B981")
                    else:
                        cell.font = Font(color="EF4444")
            elif col == 20:
                cell.font = Font(color="94A3B8") if r.get("status") == "Expired" else Font(color="10B981")

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(14, len(headers[col-1]) + 4)

    # --- Errors ---
    if errors:
        ws3 = wb.create_sheet("Errors")
        for col, h in enumerate(["Row", "Transaction Ref", "Error"], 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for i, e in enumerate(errors, 2):
            ws3.cell(row=i, column=1, value=e["row"])
            ws3.cell(row=i, column=2, value=e["ref"])
            ws3.cell(row=i, column=3, value=e["errors"]).font = Font(color="EF4444")
        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 25
        ws3.column_dimensions["C"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def generate_template():
    wb = Workbook()

    # Deals sheet
    ws = wb.active
    ws.title = "Deals"
    dh = [
        "Transaction Ref no", "Client Name", "Counterparty A",
        "Buy Contract", "Counterparty B", "Sell Contract",
        "Transaction Date / Trade Date", "Effective Date",
        "Valuation / Reporting Date", "Type of Contract",
        "Spot", "Strike Rate", "Currency PaiR",
        "Notional Currency 1", "Buy /Sell", "Notional CCY 2", "Buy /Sell",
        "Maturity Date", "Delivery Start Date", "Delivery End Date",
    ]
    for col, h in enumerate(dh, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sample = [
        "L01SFWD000130", "EXL India", "EXL India", "Buy", "BOA", "Sell",
        "07-11-2022", "07-11-2022", "31-03-2025", "Forward",
        85.47, 86.88, "USDINR", 1000000, "Sell", 8687750, "Buy", "29-07-2025",
        "", "",
    ]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)
    for col in range(1, len(dh) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Forward sheet
    ws2 = wb.create_sheet("Forward")
    for col, h in enumerate(["T", "Days", "Date", "Forwards Mid"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate([("SP",0,"4/3/2025",85.47),("1M",32,"5/5/2025",85.71),("3M",91,"7/3/2025",86.02),("6M",183,"10/3/2025",86.46),("1Y",365,"4/3/2026",87.37)], 2):
        for col, val in enumerate(row, 1):
            ws2.cell(row=i, column=col, value=val)

    # Discount sheet
    ws3 = wb.create_sheet("Discount")
    for col, h in enumerate(["Date", "Term", "Market Rate", "Zero Rate", "Discount"], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate([("4/30/2025","27 DY",7.64,7.61,0.9938),("7/31/2025","119 DY",6.84,6.79,0.9776),("3/31/2026","362 DY",6.23,6.14,0.9404),("4/5/2027","2 YR",6.26,6.17,0.8832)], 2):
        for col, val in enumerate(row, 1):
            ws3.cell(row=i, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/bulk-upload")
async def bulk_upload(
    file: UploadFile = File(...),
    domestic_rate: float = Query(0.065, description="Fallback domestic rate"),
    foreign_rate: float = Query(0.045, description="Fallback foreign rate"),
    interpolation: str = Query("linear", description="Interpolation method: 'linear' or 'cubic_spline'"),
):
    """
    Upload Excel with FX forward deals. If Forward + Discount sheets
    are present, uses curve-based pricing. Otherwise uses flat rates.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files accepted")

    if interpolation not in ("linear", "cubic_spline"):
        raise HTTPException(400, "interpolation must be 'linear' or 'cubic_spline'")

    try:
        contents = await file.read()
        deals, parse_errors, curves = parse_upload(contents)

        if not deals and parse_errors:
            raise HTTPException(400, f"No valid deals found. Errors: {parse_errors}")

        if curves and curves.has_curves():
            logger.info(f"Curve-based pricing (interpolation={interpolation})")
            results, pricing_errors = price_deals_with_curves(deals, curves, interpolation)
            method = "curve"
        else:
            logger.info("Flat-rate pricing (no curves)")
            results, pricing_errors = price_deals_flat(deals, domestic_rate, foreign_rate)
            method = "flat"

        result_bytes = generate_results_excel(results, parse_errors + pricing_errors, file.filename, method)

        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Optima_BulkResults_{file.filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Bulk upload failed")
        raise HTTPException(500, f"Bulk upload failed: {e}")


@router.get("/bulk-template")
async def download_template():
    """Download template with Deals + Forward + Discount sheets."""
    return StreamingResponse(
        io.BytesIO(generate_template()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Optima_BulkUpload_Template.xlsx"'},
    )
